from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, send_file
from flask_login import login_user, logout_user, login_required, current_user
from urllib.parse import urlparse
from models import db, User, StageAccessRequest
from forms import LoginForm, RegistrationForm, AdminApprovalForm, UserSearchForm, EditProfileForm
from datetime import datetime
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, Reference
from werkzeug.utils import secure_filename


auth = Blueprint('auth', __name__)

@auth.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('home'))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user is None or not user.check_password(form.password.data):
            flash('Invalid username or password', 'error')
            return redirect(url_for('auth.login'))
        
        if not user.can_login():
            if not user.is_approved and not user.is_admin():
                flash('Your account is pending approval. Please wait for admin approval.', 'warning')
            elif not user.is_active:
                flash('Your account has been suspended. Please contact admin.', 'error')
            return redirect(url_for('auth.login'))
        
        login_user(user, remember=form.remember_me.data)
        next_page = request.args.get('next')
        if not next_page or urlparse(next_page).netloc != '':
            next_page = url_for('home')
        
        flash(f'Welcome back, {user.full_name}!', 'success')
        return redirect(next_page)
    
    return render_template('auth/login.html', title='Login', form=form, page_title='Login - Daiva Anughara')

@auth.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('home'))
    
    form = RegistrationForm()
    if form.validate_on_submit():
        # Check for existing username
        existing_user = User.query.filter_by(username=form.username.data).first()
        if existing_user:
            flash('Username already taken. Please choose a different one.', 'error')
            return render_template('auth/register.html', title='Register', form=form)
        
        # Check for existing email
        existing_email = User.query.filter_by(email=form.email.data).first()
        if existing_email:
            flash('Email already registered. Please use a different one.', 'error')
            return render_template('auth/register.html', title='Register', form=form)
        
        # Handle profile picture upload
        profile_picture_path = None
        if form.profile_picture.data:
            file = form.profile_picture.data
            if file and file.filename:
                # Create uploads directory if it doesn't exist
                upload_dir = os.path.join(os.getcwd(), 'static', 'uploads', 'profiles')
                os.makedirs(upload_dir, exist_ok=True)
                
                # Generate secure filename
                filename = secure_filename(file.filename)
                # Add timestamp to make filename unique
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                name, ext = os.path.splitext(filename)
                filename = f"{name}_{timestamp}{ext}"
                
                # Save file
                file_path = os.path.join(upload_dir, filename)
                file.save(file_path)
                profile_picture_path = f"uploads/profiles/{filename}"
        
        user = User(
            username=form.username.data,
            email=form.email.data,
            full_name=form.full_name.data,
            phone=form.phone.data,
            address=form.address.data,
            practice_level=form.practice_level.data,
            purpose=form.purpose.data,
            profile_picture=profile_picture_path
        )
        user.set_password(form.password.data)
        
        db.session.add(user)
        db.session.commit()
        
        flash('Registration successful! Please wait for admin approval before you can login.', 'success')
        return redirect(url_for('auth.login'))
    
    return render_template('auth/register.html', title='Register', form=form, page_title='Register - Daiva Anughara')

@auth.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out successfully.', 'info')
    return redirect(url_for('home'))

@auth.route('/admin/users')
@login_required
def admin_users():
    if not current_user.is_admin():
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('home'))
    
    search_form = UserSearchForm()
    approval_form = AdminApprovalForm()
    
    # Get search parameters
    search_term = request.args.get('search_term', '')
    status_filter = request.args.get('status_filter', 'all')
    role_filter = request.args.get('role_filter', 'all')
    
    # Build query
    query = User.query
    
    if search_term:
        query = query.filter(
            db.or_(
                User.username.contains(search_term),
                User.email.contains(search_term),
                User.full_name.contains(search_term)
            )
        )
    
    if status_filter != 'all':
        if status_filter == 'pending':
            query = query.filter_by(is_approved=False, is_active=True)
        elif status_filter == 'approved':
            query = query.filter_by(is_approved=True, is_active=True)
        elif status_filter == 'rejected':
            query = query.filter_by(is_approved=False, is_active=False)
        elif status_filter == 'suspended':
            query = query.filter_by(is_active=False)
    
    if role_filter != 'all':
        query = query.filter_by(role=role_filter)
    
    users = query.order_by(User.created_at.desc()).all()
    
    # Get all pending stage access requests
    pending_requests = StageAccessRequest.query.filter_by(status='pending').order_by(StageAccessRequest.requested_at.desc()).all()
    
    return render_template('admin/users.html', 
                         title='User Management',
                         page_title='User Management - Daiva Anughara',
                         users=users,
                         pending_requests=pending_requests,
                         search_form=search_form,
                         approval_form=approval_form)

@auth.route('/admin/approve_user', methods=['POST'])
@login_required
def approve_user():
    if not current_user.is_admin():
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    form = AdminApprovalForm()
    if form.validate_on_submit():
        user_id = form.user_id.data
        action = form.action.data
        notes = form.notes.data
        
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        
        if action == 'approve':
            user.is_approved = True
            user.approved_at = datetime.utcnow()
            user.approved_by = current_user.id
            user.is_active = True
            message = f'User {user.username} has been approved successfully.'
        elif action == 'reject':
            user.is_approved = False
            user.is_active = False
            message = f'User {user.username} has been rejected.'
        elif action == 'suspend':
            user.is_active = False
            message = f'User {user.username} has been suspended.'
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': message})
    
    return jsonify({'success': False, 'message': 'Invalid form data'}), 400

@auth.route('/admin/user/<int:user_id>')
@login_required
def admin_user_detail(user_id):
    if not current_user.is_admin():
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('home'))
    
    user = User.query.get_or_404(user_id)
    # Get pending stage access requests for this user
    pending_requests = StageAccessRequest.query.filter_by(
        user_id=user_id,
        status='pending'
    ).order_by(StageAccessRequest.requested_at.desc()).all()
    
    return render_template('admin/user_detail.html', 
                         title='User Detail', 
                         page_title='User Detail - Daiva Anughara', 
                         user=user,
                         pending_requests=pending_requests)

@auth.route('/admin/user/<int:user_id>/stage-access', methods=['POST'])
@login_required
def update_stage_access(user_id):
    if not current_user.is_admin():
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('auth.admin_users'))

    user = User.query.get_or_404(user_id)

    # Get stage access updates from form (Bhairava stages 1-6)
    mandala_2_access = request.form.get('mandala_2_access') == 'on'
    mandala_3_access = request.form.get('mandala_3_access') == 'on'
    rudraksha_8_mukhi_access = request.form.get('rudraksha_8_mukhi_access') == 'on'
    rudraksha_11_mukhi_access = request.form.get('rudraksha_11_mukhi_access') == 'on'
    rudraksha_14_mukhi_access = request.form.get('rudraksha_14_mukhi_access') == 'on'
    pratham_charana_diksha_access = request.form.get('pratham_charana_diksha_access') == 'on'
    dutiya_charana_access = request.form.get('dutiya_charana_access') == 'on'
    tritiya_charana_access = request.form.get('tritiya_charana_access') == 'on'

    # Get Devi Mandala access updates (Kamakhya Sadhana)
    devi_mandala_2_access = request.form.get('devi_mandala_2_access') == 'on'
    devi_mandala_3_access = request.form.get('devi_mandala_3_access') == 'on'

    # Update Bhairava stage access
    user.mandala_2_access = mandala_2_access
    user.mandala_3_access = mandala_3_access
    user.rudraksha_8_mukhi_access = rudraksha_8_mukhi_access
    user.rudraksha_11_mukhi_access = rudraksha_11_mukhi_access
    user.rudraksha_14_mukhi_access = rudraksha_14_mukhi_access
    user.pratham_charana_diksha_access = pratham_charana_diksha_access
    user.dutiya_charana_access = dutiya_charana_access
    user.tritiya_charana_access = tritiya_charana_access

    # Update Devi Mandala access
    user.devi_mandala_2_access = devi_mandala_2_access
    user.devi_mandala_3_access = devi_mandala_3_access

    # Check if user should start the next available stage
    next_stage = user.get_next_required_stage()
    if next_stage and not getattr(user, f'mandala_{next_stage}_started_at' if next_stage <= 3 else f'rudraksha_{5 if next_stage == 4 else 11 if next_stage == 5 else 14}_mukhi_started_at', None):
        user.start_stage(next_stage)

    db.session.commit()

    flash(f'Stage access updated for {user.username}', 'success')
    return redirect(url_for('auth.admin_user_detail', user_id=user_id))

@auth.route('/admin/user/<int:user_id>/complete-stage', methods=['POST'])
@login_required
def complete_user_stage(user_id):
    if not current_user.is_admin():
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('auth.admin_users'))

    user = User.query.get_or_404(user_id)
    stage_number = int(request.form.get('stage_number', 0))

    if stage_number < 1 or stage_number > 6:
        flash('Invalid stage number', 'error')
        return redirect(url_for('auth.admin_user_detail', user_id=user_id))

    # Check if user has access to this stage
    if not user.has_mandala_access(stage_number):
        flash('User does not have access to this stage', 'error')
        return redirect(url_for('auth.admin_user_detail', user_id=user_id))

    # Check if stage is already completed
    if user.is_stage_completed(stage_number):
        flash('Stage is already completed', 'warning')
        return redirect(url_for('auth.admin_user_detail', user_id=user_id))

    # Complete the stage
    user.complete_stage(stage_number)

    # Grant access to next stage if this is not the last stage
    if stage_number < 6:
        next_stage = stage_number + 1
        access_fields = {
            2: 'mandala_2_access',
            3: 'mandala_3_access',
            4: 'rudraksha_8_mukhi_access',
            5: 'rudraksha_11_mukhi_access',
            6: 'rudraksha_14_mukhi_access'
        }

        if next_stage in access_fields:
            setattr(user, access_fields[next_stage], True)
            user.start_stage(next_stage)

    db.session.commit()

    flash(f'Stage {stage_number} completed for {user.username}', 'success')
    return redirect(url_for('auth.admin_user_detail', user_id=user_id))

@auth.route('/admin/user/<int:user_id>/reset-stage', methods=['POST'])
@login_required
def reset_user_stage(user_id):
    if not current_user.is_admin():
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('auth.admin_users'))

    user = User.query.get_or_404(user_id)
    stage_number = int(request.form.get('stage_number', 0))

    if stage_number < 1 or stage_number > 6:
        flash('Invalid stage number', 'error')
        return redirect(url_for('auth.admin_user_detail', user_id=user_id))

    # Reset stage completion and dates
    completion_fields = {
        1: 'mandala_1_completed_at',
        2: 'mandala_2_completed_at',
        3: 'mandala_3_completed_at',
        4: 'rudraksha_8_mukhi_completed_at',
        5: 'rudraksha_11_mukhi_completed_at',
        6: 'rudraksha_14_mukhi_completed_at'
    }

    start_fields = {
        1: 'mandala_1_started_at',
        2: 'mandala_2_started_at',
        3: 'mandala_3_started_at',
        4: 'rudraksha_8_mukhi_started_at',
        5: 'rudraksha_11_mukhi_started_at',
        6: 'rudraksha_14_mukhi_started_at'
    }

    if stage_number in completion_fields:
        setattr(user, completion_fields[stage_number], None)
        setattr(user, start_fields[stage_number], None)

    # If resetting a stage, also reset all subsequent stages
    for i in range(stage_number + 1, 7):
        if i in completion_fields:
            setattr(user, completion_fields[i], None)
            setattr(user, start_fields[i], None)

        # Remove access to subsequent stages
        if i == 2:
            user.mandala_2_access = False
        elif i == 3:
            user.mandala_3_access = False
        elif i == 4:
            user.rudraksha_8_mukhi_access = False
        elif i == 5:
            user.rudraksha_11_mukhi_access = False
        elif i == 6:
            user.rudraksha_14_mukhi_access = False

    db.session.commit()

    flash(f'Stage {stage_number} and subsequent stages reset for {user.username}', 'success')
    return redirect(url_for('auth.admin_user_detail', user_id=user_id))

@auth.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    form = EditProfileForm()
    
    if form.validate_on_submit():
        current_user.address = form.address.data
        current_user.purpose = form.purpose.data
        current_user.phone = form.phone.data
        
        try:
            db.session.commit()
            flash('Profile updated successfully!', 'success')
            return redirect(url_for('auth.profile'))
        except Exception as e:
            db.session.rollback()
            flash('Error updating profile. Please try again.', 'error')
    
    # Pre-fill form
    if request.method == 'GET':
        form.address.data = current_user.address
        form.purpose.data = current_user.purpose
        form.phone.data = current_user.phone

    return render_template('auth/profile.html', 
                         title='Profile', 
                         page_title='Profile - Daiva Anughara',
                         form=form)

@auth.route('/profile/edit', methods=['GET', 'POST'])
@login_required
def edit_profile():
    # This would be implemented for users to edit their own profile
    pass

@auth.route('/admin/user/<int:user_id>/update-profile-picture', methods=['POST'])
@login_required
def update_profile_picture(user_id):
    """Admin endpoint to update a user's profile picture"""
    if not current_user.is_admin():
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('home'))
    
    user = User.query.get_or_404(user_id)
    
    if 'profile_picture' not in request.files:
        flash('No file selected.', 'error')
        return redirect(url_for('auth.admin_user_detail', user_id=user_id))
    
    file = request.files['profile_picture']
    
    if file.filename == '':
        flash('No file selected.', 'error')
        return redirect(url_for('auth.admin_user_detail', user_id=user_id))
    
    if file:
        # Validate file extension
        allowed_extensions = {'jpg', 'jpeg', 'png', 'gif'}
        filename = secure_filename(file.filename)
        ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
        
        if ext not in allowed_extensions:
            flash('Invalid file type. Only JPG, PNG, and GIF are allowed.', 'error')
            return redirect(url_for('auth.admin_user_detail', user_id=user_id))
        
        # Create uploads directory if it doesn't exist
        if os.getenv('VERCEL'):
            upload_dir = os.path.join('/tmp', 'static', 'uploads', 'profiles')
        else:
            upload_dir = os.path.join(os.getcwd(), 'static', 'uploads', 'profiles')
            
        try:
            os.makedirs(upload_dir, exist_ok=True)
        except Exception as e:
            flash(f"Upload failed: Read-only serverless filesystem. {e}", 'error')
            return redirect(url_for('auth.admin_user_detail', user_id=user_id))
        
        # Generate secure filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        name, _ = os.path.splitext(filename)
        new_filename = f"{name}_{user.id}_{timestamp}.{ext}"
        
        # Save file
        file_path = os.path.join(upload_dir, new_filename)
        file.save(file_path)
        
        # Delete old profile picture if exists
        if user.profile_picture:
            old_path = os.path.join(os.getcwd(), 'static', user.profile_picture)
            if os.path.exists(old_path):
                try:
                    os.remove(old_path)
                except Exception as e:
                    print(f"Warning: Could not delete old profile picture: {e}")
        
        # Update user's profile picture path
        user.profile_picture = f"uploads/profiles/{new_filename}"
        db.session.commit()
        
        flash(f'Profile picture updated successfully for {user.username}!', 'success')
    
    return redirect(url_for('auth.admin_user_detail', user_id=user_id))

def generate_user_report():
    users = User.query.all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "User Report"

    # Headers
    headers = [
        "User ID", "Username", "Full Name", "Email", "Role", "Status",
        "Join Date", "Approval Date", "Approved for Mandala 2", "Approved for Mandala 3",
        "Days on Mandala 1", "Days on Mandala 2", "Days on Mandala 3"
    ]
    sheet.append(headers)

    # Style for headers
    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data
    for user in users:
        status = "Admin" if user.is_admin() else ("Approved" if user.is_approved else ("Suspended" if not user.is_active else "Pending"))
        
        days_on_mandala_1 = (datetime.utcnow() - user.created_at).days if user.created_at else 0
        days_on_mandala_2 = 0
        days_on_mandala_3 = 0

        if user.mandala_2_access and user.approved_at:
            days_on_mandala_2 = (datetime.utcnow() - user.approved_at).days
        
        if user.mandala_3_access and user.approved_at: # Assuming mandala 3 access is granted at the same time or after mandala 2
            days_on_mandala_3 = (datetime.utcnow() - user.approved_at).days


        sheet.append([
            user.id,
            user.username,
            user.full_name,
            user.email,
            user.role,
            status,
            user.created_at.strftime("%Y-%m-%d") if user.created_at else "",
            user.approved_at.strftime("%Y-%m-%d") if user.approved_at else "",
            "Yes" if user.mandala_2_access else "No",
            "Yes" if user.mandala_3_access else "No",
            days_on_mandala_1,
            days_on_mandala_2,
            days_on_mandala_3
        ])

    # KPIs
    sheet.cell(row=1, column=15, value="KPIs").font = header_font
    
    total_users = len(users)
    approved_users = len([u for u in users if u.is_approved])
    pending_users = len([u for u in users if not u.is_approved and u.is_active])
    
    sheet.cell(row=2, column=15, value="Total Users")
    sheet.cell(row=2, column=16, value=total_users)
    sheet.cell(row=3, column=15, value="Approved Users")
    sheet.cell(row=3, column=16, value=approved_users)
    sheet.cell(row=4, column=15, value="Pending Users")
    sheet.cell(row=4, column=16, value=pending_users)

    # Chart
    chart_sheet = workbook.create_sheet(title="User Status Chart")
    chart_data = [
        ['Status', 'Count'],
        ['Approved', approved_users],
        ['Pending', pending_users],
        ['Suspended', total_users - approved_users - pending_users]
    ]
    for row in chart_data:
        chart_sheet.append(row)
    
    chart = BarChart()
    chart.title = "User Status Distribution"
    chart.y_axis.title = "Number of Users"
    chart.x_axis.title = "Status"
    
    data = Reference(chart_sheet, min_col=2, min_row=1, max_row=4, max_col=2)
    cats = Reference(chart_sheet, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart_sheet.add_chart(chart, "E5")


    # Save to a BytesIO object
    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    return excel_file

@auth.route('/admin/users/report')
@login_required
def user_report():
    if not current_user.is_admin():
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('home'))

    excel_file = generate_user_report()

    return send_file(
        excel_file,
        as_attachment=True,
        download_name='user_report.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@auth.route('/admin/users/kpi_data')
@login_required
def user_kpi_data():
    if not current_user.is_admin():
        return jsonify({'error': 'Access denied'}), 403

    users = User.query.all()

    total_users = len(users)
    approved_users = len([u for u in users if u.is_approved and u.is_active])
    pending_users = len([u for u in users if not u.is_approved and u.is_active])
    suspended_users = len([u for u in users if not u.is_active and not u.is_admin()])

    mandala_2_users = len([u for u in users if u.mandala_2_access])
    mandala_3_users = len([u for u in users if u.mandala_3_access])

    # Calculate average days to approval
    approved_users_with_approval_date = [u for u in users if u.is_approved and u.approved_at and u.created_at]
    if approved_users_with_approval_date:
        total_days_to_approval = sum([(u.approved_at - u.created_at).days for u in approved_users_with_approval_date])
        average_days_to_approval = total_days_to_approval / len(approved_users_with_approval_date)
    else:
        average_days_to_approval = 0

    data = {
        'kpis': {
            'total_users': total_users,
            'approved_users': approved_users,
            'pending_users': pending_users,
            'suspended_users': suspended_users,
            'average_days_to_approval': round(average_days_to_approval, 2)
        },
        'charts': {
            'user_status_distribution': {
                'labels': ['Approved', 'Pending', 'Suspended'],
                'data': [approved_users, pending_users, suspended_users]
            },
            'mandala_access_distribution': {
                'labels': ['Mandala 2 Access', 'Mandala 3 Access'],
                'data': [mandala_2_users, mandala_3_users]
            }
        }
    }
    return jsonify(data)

@auth.route('/request_stage_access', methods=['POST'])
@login_required
def request_stage_access():
    """Allow users to request access for a locked stage"""
    try:
        stage_number = request.form.get('stage_number', type=int)
        
        if not stage_number or stage_number < 1 or stage_number > 9:
            return jsonify({'success': False, 'message': 'Invalid stage number'}), 400
        
        # Check if user already has access to this stage
        if current_user.has_mandala_access(stage_number):
            return jsonify({'success': False, 'message': 'You already have access to this stage'}), 400
        
        # Check if there's already a pending request for this stage
        existing_request = StageAccessRequest.query.filter_by(
            user_id=current_user.id,
            stage_number=stage_number,
            status='pending'
        ).first()
        
        if existing_request:
            return jsonify({'success': False, 'message': 'You already have a pending request for this stage'}), 400
        
        # Create new request
        new_request = StageAccessRequest(
            user_id=current_user.id,
            stage_number=stage_number,
            status='pending'
        )
        
        db.session.add(new_request)
        db.session.commit()
        
        stage_names = {
            1: 'Mandala 1',
            2: 'Mandala 2',
            3: 'Mandala 3',
            4: 'Rudraksha 8 Mukhi',
            5: 'Rudraksha 11 Mukhi',
            6: 'Rudraksha 14 Mukhi',
            7: 'Pratham Charana Diksha',
            8: 'Dutiya Charana',
            9: 'Tritiya Charana'
        }
        stage_name = stage_names.get(stage_number, f'Stage {stage_number}')
        
        return jsonify({
            'success': True,
            'message': f'Access request for {stage_name} has been sent to administrators'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': 'An error occurred. Please try again.'}), 500

@auth.route('/request_devi_stage_access', methods=['POST'])
@login_required
def request_devi_stage_access():
    """Allow users to request access for a locked Devi Mandala stage"""
    try:
        mandala_number = request.form.get('mandala_number', type=int)
        
        if not mandala_number or mandala_number < 1 or mandala_number > 3:
            return jsonify({'success': False, 'message': 'Invalid mandala number'}), 400
        
        # Check if user is approved
        if not current_user.is_approved:
            return jsonify({'success': False, 'message': 'Your account needs to be approved first'}), 400
        
        # Check if there's already a pending request for this Devi mandala
        # Using stage_number 101, 102, 103 for Devi mandalas to distinguish from Bhairava stages
        devi_stage_number = 100 + mandala_number
        
        existing_request = StageAccessRequest.query.filter_by(
            user_id=current_user.id,
            stage_number=devi_stage_number,
            status='pending'
        ).first()
        
        if existing_request:
            return jsonify({'success': False, 'message': 'You already have a pending request for this Devi Mandala'}), 400
        
        # Create new request
        new_request = StageAccessRequest(
            user_id=current_user.id,
            stage_number=devi_stage_number,
            status='pending'
        )
        
        db.session.add(new_request)
        db.session.commit()
        
        mandala_days = {1: 33, 2: 66, 3: 99}
        
        return jsonify({
            'success': True,
            'message': f'Access request for Devi Mandala {mandala_number} ({mandala_days.get(mandala_number, 33)} days) has been sent to administrators'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': 'An error occurred. Please try again.'}), 500

@auth.route('/admin/approve_stage_request/<int:request_id>', methods=['POST'])
@login_required
def approve_stage_request(request_id):
    """Admin endpoint to approve a stage access request"""
    if not current_user.is_admin():
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    try:
        access_request = StageAccessRequest.query.get_or_404(request_id)
        
        if access_request.status != 'pending':
            return jsonify({'success': False, 'message': 'Request has already been processed'}), 400
        
        action = request.form.get('action', 'approve')  # 'approve' or 'reject'
        
        if action == 'approve':
            # Grant access to the stage
            user = access_request.user
            
            # Grant access based on stage number
            if access_request.stage_number == 2:
                user.mandala_2_access = True
            elif access_request.stage_number == 3:
                user.mandala_3_access = True
            elif access_request.stage_number == 4:
                user.rudraksha_8_mukhi_access = True
            elif access_request.stage_number == 5:
                user.rudraksha_11_mukhi_access = True
            elif access_request.stage_number == 6:
                user.rudraksha_14_mukhi_access = True
            elif access_request.stage_number == 7:
                user.pratham_charana_diksha_access = True
            elif access_request.stage_number == 8:
                user.dutiya_charana_access = True
            elif access_request.stage_number == 9:
                user.tritiya_charana_access = True
            # Devi Mandala stages (Devi Padathi - Kamakhya Sadhana)
            elif access_request.stage_number == 101:
                user.devi_mandala_1_access = True
            elif access_request.stage_number == 102:
                user.devi_mandala_2_access = True
            elif access_request.stage_number == 103:
                user.devi_mandala_3_access = True
            
            # Start the stage if it's the next one
            next_stage = user.get_next_required_stage()
            if next_stage == access_request.stage_number:
                user.start_stage(access_request.stage_number)
            
            access_request.status = 'approved'
            access_request.reviewed_at = datetime.utcnow()
            access_request.reviewed_by = current_user.id
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': f'Access to {access_request.get_stage_name()} has been granted to {user.username}'
            })
        else:  # reject
            access_request.status = 'rejected'
            access_request.reviewed_at = datetime.utcnow()
            access_request.reviewed_by = current_user.id
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': f'Access request for {access_request.get_stage_name()} has been rejected'
            })
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': 'An error occurred. Please try again.'}), 500

@auth.route('/admin/user/<int:user_id>/change-password', methods=['POST'])
@login_required
def admin_change_password(user_id):
    """Admin endpoint to change a user's password"""
    if not current_user.is_admin():
        return jsonify({'success': False, 'message': 'Access denied. Admin privileges required.'}), 403
    
    try:
        user = User.query.get_or_404(user_id)
        
        # Prevent changing admin passwords (security measure)
        if user.is_admin() and user.id != current_user.id:
            return jsonify({'success': False, 'message': 'Cannot change password of another admin'}), 403
        
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        
        # Validation
        if not new_password:
            return jsonify({'success': False, 'message': 'Password is required'}), 400
        
        if len(new_password) < 6:
            return jsonify({'success': False, 'message': 'Password must be at least 6 characters long'}), 400
        
        if new_password != confirm_password:
            return jsonify({'success': False, 'message': 'Passwords do not match'}), 400
        
        # Update password
        user.set_password(new_password)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Password successfully changed for {user.username}'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': 'An error occurred. Please try again.'}), 500